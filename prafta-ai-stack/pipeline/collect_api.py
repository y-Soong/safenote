# -*- coding: utf-8 -*-
"""
오픈API 수집기 — 라이선스 통과(license_checked=TRUE)한 feed_type=API 자료를
페이지네이션으로 전량 호출해 corpus/raw/ 에 JSON 스냅샷으로 저장한다.

역할은 "받은 원본을 무변경 보관"까지다. 마스킹·정제·청킹은 이후 단계(10_preprocess 등).
따라서 여기서는 응답 item 을 가공하지 않고 그대로 스냅샷한다(raw/ 는 .gitignore 제외).

전제:
  - corpus/.env 에 DATA_GO_KR_SERVICE_KEY=... (실제 키, git 제외)
  - corpus/registry/sources.xlsx 에 자료 등록 + license_checked=TRUE

사용:
  python pipeline/collect_api.py --probe            # 연결/키 확인(totalCount만, 저장 안 함)
  python pipeline/collect_api.py                    # 라이선스 통과 API 자료 전량 수집
  python pipeline/collect_api.py --source 15053339  # 특정 자료만
  python pipeline/collect_api.py --rows 500         # 페이지당 건수(기본 1000)

의존성: requests, openpyxl (Anaconda 기본 포함)
"""
import argparse
import json
import os
import sys
import time
from datetime import datetime, date

try:
    import requests
except ImportError:
    sys.exit("[실패] requests 미설치. `pip install requests` 후 재시도.")
try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("[실패] openpyxl 미설치. `pip install openpyxl` 후 재시도.")

# ── 경로(스크립트 기준으로 고정) ──
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)                       # prafta-ai-stack/
ENV_PATH = os.path.join(ROOT, "corpus", ".env")
REGISTRY_PATH = os.path.join(ROOT, "corpus", "registry", "sources.xlsx")
RAW_DIR = os.path.join(ROOT, "corpus", "raw")

LIST_COLS = {"content_fields", "mask_fields", "exclude_from_llm", "locator_fields", "meta_fields"}
OK_RESULT_CODES = {"0", "00"}                      # data.go.kr NORMAL


def load_env(path):
    """corpus/.env 파싱(KEY=VALUE). python-dotenv 의존 없이."""
    env = {}
    if not os.path.exists(path):
        sys.exit(f"[실패] .env 없음: {path}\n  → .env.example 을 .env 로 복사하고 키를 넣으세요.")
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def load_registry(path):
    """sources.xlsx -> list[dict]. 리스트 컬럼은 쉼표 분리."""
    if not os.path.exists(path):
        sys.exit(f"[실패] registry 없음: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["sources"] if "sources" in wb.sheetnames else wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(it)]
    rows = []
    for raw in it:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        d = {}
        for h, v in zip(headers, raw):
            if not h:
                continue
            if h in LIST_COLS:
                d[h] = [s.strip() for s in str(v).split(",") if s and s.strip()] if v not in (None, "") else []
            else:
                d[h] = ("" if v is None else v)
        rows.append(d)
    wb.close()
    return rows


def is_checked(v):
    """license_checked 해석: 엑셀 bool True 또는 문자열 'TRUE'/'Y'/'1'."""
    if v is True:
        return True
    return str(v).strip().upper() in {"TRUE", "Y", "1"}


def http_get_json(url, params, retries=3, timeout=30):
    """단순 재시도 포함 GET → JSON dict."""
    last = None
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(url, params=params, timeout=timeout)
            r.raise_for_status()
            return r.json()
        except Exception as e:                     # noqa: BLE001 (수집 견고성 우선)
            last = e
            if attempt < retries:
                time.sleep(1.5 * attempt)
    raise RuntimeError(f"GET 실패({retries}회): {last}")


def parse_response(js):
    """data.go.kr 표준 응답에서 (resultCode, totalCount, items) 추출."""
    resp = js.get("response", js)
    header = resp.get("header", {}) or {}
    body = resp.get("body", {}) or {}
    code = str(header.get("resultCode", "")).strip()
    total = int(str(body.get("totalCount", "0")).strip() or 0)
    items_wrap = body.get("items", "")
    item = ""
    if isinstance(items_wrap, dict):
        item = items_wrap.get("item", "")
    # item 은 dict(1건)·list(N건)·""(0건) 가능 → list 로 정규화
    if isinstance(item, dict):
        items = [item]
    elif isinstance(item, list):
        items = item
    else:
        items = []
    return code, total, items


def collect_source(src, service_key, rows_per_page, probe):
    sid = str(src.get("source_id", "")).strip()
    endpoint = str(src.get("api_endpoint", "")).strip()
    if not endpoint:
        print(f"  [건너뜀] {sid}: api_endpoint 없음")
        return None

    base = {"ServiceKey": service_key, "type": "json"}

    # 소스별 고정 파라미터 병합(registry api_fixed_params, 예: KOSHA callApiId=1060).
    # 형식: "k=v" 또는 "k1=v1&k2=v2". 일부 제공기관(KOSHA B552468)은 라우팅용 필수 고정값을
    # 요구하며, 누락 시 백엔드가 500 "Unexpected errors" 로 크래시한다.
    fixed = str(src.get("api_fixed_params", "")).strip()
    if fixed:
        for kv in fixed.split("&"):
            if "=" in kv:
                k, v = kv.split("=", 1)
                base[k.strip()] = v.strip()

    # 1) 첫 호출로 totalCount 확인
    first = http_get_json(endpoint, {**base, "pageNo": 1, "numOfRows": 1})
    code, total, _ = parse_response(first)
    if code not in OK_RESULT_CODES:
        print(f"  [실패] {sid}: resultCode={code} (인증키/승인 상태 확인 필요)")
        return None
    print(f"  totalCount = {total:,} 건")
    if probe:
        return {"source_id": sid, "total_count": total, "probe": True}
    if total == 0:
        print(f"  [주의] {sid}: 0건 → 저장 생략")
        return None

    # 2) 페이지네이션 전량 수집
    #    일부 API는 numOfRows 상한(예: 20)을 둬서 요청값보다 적게 반환한다.
    #    → 1페이지 실제 반환수로 stride(페이지 크기)를 확정해 pageNo 계산을 일치시킨다.
    requested = rows_per_page
    all_items = []
    stride = requested
    p = 1
    max_pages = total + 5                            # 안전 상한(최악: 페이지당 1건)
    while len(all_items) < total and p <= max_pages:
        js = http_get_json(endpoint, {**base, "pageNo": p, "numOfRows": stride})
        code, _, items = parse_response(js)
        if code not in OK_RESULT_CODES:
            print(f"  [실패] {sid} p{p}: resultCode={code} — 여기까지 수집분으로 중단")
            break
        if not items:
            print(f"  [주의] {sid} p{p}: 빈 페이지 — 수집 종료(누적 {len(all_items):,}/{total:,})")
            break
        all_items.extend(items)
        if p == 1 and len(items) < requested:       # API 상한 감지 → stride 확정
            stride = len(items)
        print(f"    p{p}  누적 {len(all_items):,}/{total:,}", flush=True)
        p += 1

    # 3) raw 스냅샷 저장(무변경 items + 수집 메타)
    os.makedirs(RAW_DIR, exist_ok=True)
    today = date.today().strftime("%Y%m%d")
    fname = str(src.get("file", f"raw/{sid}_{today}.json"))
    fname = os.path.basename(fname).replace("YYYYMMDD", today)
    out_path = os.path.join(RAW_DIR, fname)
    snapshot = {
        "_collection": {
            "source_id": sid,
            "source_name": src.get("source_name", ""),
            "endpoint": endpoint,
            "collected_at": datetime.now().isoformat(timespec="seconds"),
            "total_count": total,
            "collected_count": len(all_items),
            "pages": p - 1,
            "rows_per_page": stride,
        },
        "items": all_items,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2)
    print(f"  [저장] {out_path}  ({len(all_items):,}건)")
    return {"source_id": sid, "total_count": total, "collected": len(all_items), "path": out_path}


def main():
    ap = argparse.ArgumentParser(description="오픈API 수집기(raw 스냅샷)")
    ap.add_argument("--probe", action="store_true", help="totalCount만 확인(저장 안 함)")
    ap.add_argument("--source", help="특정 source_id만")
    ap.add_argument("--rows", type=int, default=1000, help="페이지당 건수(기본 1000)")
    args = ap.parse_args()

    env = load_env(ENV_PATH)
    service_key = env.get("DATA_GO_KR_SERVICE_KEY", "")
    if not service_key or service_key.startswith("여기에"):
        sys.exit("[실패] DATA_GO_KR_SERVICE_KEY 미설정. corpus/.env 확인.")

    registry = load_registry(REGISTRY_PATH)

    targets = []
    for src in registry:
        sid = str(src.get("source_id", "")).strip()
        if args.source and sid != str(args.source):
            continue
        if str(src.get("feed_type", "")).strip().upper() != "API":
            continue
        if not is_checked(src.get("license_checked")):
            print(f"[게이트] {sid}: license_checked 아님 → 건너뜀(하드가드#1)")
            continue
        targets.append(src)

    if not targets:
        sys.exit("[안내] 수집 대상(API + license_checked=TRUE) 없음.")

    print(f"== 수집 시작 ({'PROBE' if args.probe else '전량'}) · 대상 {len(targets)}건 ==")
    results = []
    for src in targets:
        print(f"\n[{src.get('source_id')}] {src.get('source_name')}")
        try:
            r = collect_source(src, service_key, args.rows, args.probe)
            if r:
                results.append(r)
        except Exception as e:                     # noqa: BLE001
            print(f"  [오류] {src.get('source_id')}: {e}")

    print("\n== 요약 ==")
    for r in results:
        if r.get("probe"):
            print(f"  {r['source_id']}: totalCount {r['total_count']:,}")
        else:
            print(f"  {r['source_id']}: {r.get('collected', 0):,}/{r['total_count']:,} 건 → {r.get('path','')}")


if __name__ == "__main__":
    main()
