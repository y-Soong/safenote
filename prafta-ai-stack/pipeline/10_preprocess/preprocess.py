# -*- coding: utf-8 -*-
"""
10_preprocess — 원본(raw) → 청크직전(processed) 변환.

레지스트리(sources.xlsx)의 preprocess 매핑을 그대로 적용한다(작업지시서 §4·§5, 결정①②③).
  - CONTENT = content_fields 결합(+ \r\n 정규화)
  - 마스킹(결정③): mask_fields 값 + 본문 스캔 익명치환, 주소 시군구 절단
  - MEASURE_TEXT: measure_field 얕으면 NULL(§4-3). HAZARD_TEXT: 전용필드 없으면 NULL(결정①)
  - META_JSON = meta_fields + exclude_from_llm(감사용·LLM 미투입). LLM 입력엔 CONTENT만.
  - CHUNK_ID(결정②) = {SOURCE_ID}_{sha256(SOURCE_ID|locator|CONTENT)[:16]}  (멱등 결정적 ID)
  - license_checked=TRUE 자료만 처리(하드가드#1)

이번 리비전은 feed_type=API(수집 스냅샷 JSON) 리더를 구현한다.
FILE(csv/xlsx/pdf) 리더는 후속 자료에서 순차 추가.

사용: python pipeline/10_preprocess/preprocess.py [--source 15053339]
"""
import argparse
import glob
import hashlib
import importlib.util
import json
import os
import re
import sys

# Windows 콘솔 cp949 크래시 방어(★반복된 함정): em dash(—) 등 일부 유니코드 문자를 print 하면
# UnicodeEncodeError 로 스크립트 전체가 죽는다(20_collect_disaster_attach.py 에서 실제로 겪음).
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

HERE = os.path.dirname(os.path.abspath(__file__))
PIPE = os.path.dirname(HERE)
ROOT = os.path.dirname(PIPE)
PROCESSED_DIR = os.path.join(ROOT, "corpus", "processed")


def _load(modname, path):
    spec = importlib.util.spec_from_file_location(modname, path)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m

_ca = _load("collect_api", os.path.join(PIPE, "collect_api.py"))
_mask = _load("masking", os.path.join(PIPE, "common", "masking.py"))

REGISTRY_PATH = _ca.REGISTRY_PATH
RAW_DIR = _ca.RAW_DIR

# MEASURE_TEXT NULL 판정(§4-3): 너무 짧거나 정크면 대책을 NULL 처리(행은 유지)
MEASURE_MIN_LEN = 12
MEASURE_JUNK = {"미입력", "해당없음", "-", "없음", "지속 치료", "지속치료", "안전교육", "n/a", "N/A"}

_WS = re.compile(r"[ \t]+")
_MULTINL = re.compile(r"\n{3,}")


def normalize_text(t):
    t = str(t or "").replace("\r\n", "\n").replace("\r", "\n")
    t = "\n".join(_WS.sub(" ", ln).strip() for ln in t.split("\n"))
    return _MULTINL.sub("\n\n", t).strip()


def is_addr_field(name):
    n = str(name)
    return n.endswith("Addr") or "주소" in n or "addr" in n.lower()


def chunk_id(sid, locator, content):
    h = hashlib.sha256(f"{sid}|{locator}|{content}".encode("utf-8")).hexdigest()[:16]
    return f"{sid}_{h}"


def measure_or_null(val):
    v = normalize_text(val)
    if len(v) < MEASURE_MIN_LEN or v in MEASURE_JUNK:
        return None
    return v


# ── 리더: feed_type/확장자별 원본 → dict 레코드 리스트 ──
def read_api(src):
    sid = str(src["source_id"]).strip()
    hits = sorted(glob.glob(os.path.join(RAW_DIR, f"{sid}_*.json")))
    if not hits:
        raise FileNotFoundError(f"수집 스냅샷 없음: {sid}_*.json (collect_api 먼저 실행)")
    snap = json.load(open(hits[-1], encoding="utf-8"))
    return snap.get("items", [])


def read_csv_file(src):
    import csv
    fname = os.path.basename(str(src.get("file", "")).strip())
    path = os.path.join(RAW_DIR, fname)
    if not os.path.exists(path):
        raise FileNotFoundError(f"원본 CSV 없음: {fname}")
    enc = str(src.get("encoding") or "utf-8").strip() or "utf-8"
    with open(path, encoding=enc, newline="") as f:
        return list(csv.DictReader(f))


def read_sif_xlsx(src):
    """SIF(15140383) 전용 리더: 아카이브(제조업등)/(건설업) 2시트 분기.
    시트마다 헤더 위치·스키마가 다르다 → 헤더행(재해개요 포함) 자동탐지 + 병합 서브헤더 흡수.
    각 레코드에 통합키 _sheet, sif_domain 부여(도메인/로케이터 정합)."""
    from openpyxl import load_workbook
    path = os.path.join(RAW_DIR, os.path.basename(str(src.get("file", "")).strip()))
    if not os.path.exists(path):
        raise FileNotFoundError(f"원본 xlsx 없음: {os.path.basename(path)}")
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    for sh in wb.sheetnames:
        if "아카이브" not in sh:            # '개요' 등 제외
            continue
        sheet_label = "건설업" if "건설" in sh else "제조업등"
        ws = wb[sh]
        rows = [list(r) if r is not None else [] for r in ws.iter_rows(values_only=True)]

        def isdigit(v):
            return v is not None and str(v).strip().isdigit()

        hidx = next((i for i, r in enumerate(rows)
                     if any(str(c or "").strip() == "재해개요" for c in r)), None)
        if hidx is None:
            continue
        # 헤더 내부 줄바꿈 제거('산재업종\n(대분류)' → '산재업종(대분류)'). 일반 공백은 보존.
        header = [re.sub(r"\s*\n\s*", "", str(c).strip()) if c is not None else "" for c in rows[hidx]]
        ycol = next((j for j, h in enumerate(header) if h == "연번"), None)

        nxt = rows[hidx + 1] if hidx + 1 < len(rows) else []
        nxt_ynum = ycol is not None and len(nxt) > ycol and isdigit(nxt[ycol])
        if nxt_ynum:                        # 다음 행이 데이터(제조업등)
            data_start = hidx + 1
        else:                               # 다음 행이 서브헤더(건설업: 공종/작업명/단위작업명)
            for j in range(len(header)):
                sv = nxt[j] if j < len(nxt) else None
                if sv is not None and str(sv).strip():
                    header[j] = re.sub(r"\s*\n\s*", "", str(sv).strip())   # 서브헤더로 대체(더 구체적)
            data_start = hidx + 2

        for r in rows[data_start:]:
            v0 = r[ycol] if ycol is not None and len(r) > ycol else None
            if not isdigit(v0):             # 연번 없는 행 제외(빈행/구분행)
                continue
            rec = {}
            for j, h in enumerate(header):
                if h:
                    rec[h] = r[j] if j < len(r) else None
            rec["_sheet"] = sheet_label
            rec["sif_domain"] = (str(rec.get("산재업종(대분류)") or "").strip()
                                 or sheet_label)
            out.append(rec)
    wb.close()
    return out


def read_pdf(src):
    """15140227 전용 리더: 발전소 기인물별 유해위험요인·감소대책 표 PDF.
    pdfplumber 표 추출(괘선 기반 셀 복원). 개요 page0 제외, 2행 헤더 스킵(순번 숫자 행만),
    기인물명 병합셀은 forward-fill. 컬럼: 순번/기인물명/유해위험요인/감소대책/가능성/중대성/위험성."""
    import warnings
    warnings.filterwarnings("ignore")
    import pdfplumber
    path = os.path.join(RAW_DIR, os.path.basename(str(src.get("file", "")).strip()))
    if not os.path.exists(path):
        raise FileNotFoundError(f"원본 PDF 없음: {os.path.basename(path)}")
    out, last = [], ""
    with pdfplumber.open(path) as pdf:
        for pi, page in enumerate(pdf.pages):
            if pi == 0:                     # 요약(개요) 페이지 제외
                continue
            for t in (page.extract_tables() or []):
                for row in t:
                    if not row:
                        continue
                    c0 = str(row[0] or "").strip()
                    if not c0.isdigit():    # 헤더/구분행 제외(순번 숫자만 데이터)
                        continue
                    agent = re.sub(r"\s*\n\s*", "", str(row[1] or "").strip()) or last
                    last = agent

                    def g(i):
                        return str(row[i]).strip() if len(row) > i and row[i] is not None else ""
                    out.append({
                        "순번": c0, "기인물명": agent,
                        "유해위험요인": g(2), "감소대책": g(3),
                        "가능성": g(4), "중대성": g(5), "위험성": g(6),
                        "_domain": "발전설비안전",
                    })
    return out


# 15121008 첨부 PDF는 여러 해에 걸쳐 수집된 자료라 템플릿이 하나가 아니다(2026-08-26 748건 표본
# 실측으로 발견). 페이지 크기로 템플릿을 감지해 서로 다른 파서를 태운다.
#   Template A: 소형 "OPS" 카드(~252x356pt, 최근 2026년 케이스). 발생원인/예방대책 좌우 2단.
#   Template B: A4 세로(~595x841pt, 과거 케이스 다수). 단일 컬럼 서술형. 헤더 표기가 문서마다
#     다르다(발생원인/재해발생원인/재해발생 원인 등) — 정확 일치 대신 공백제거+부분일치로 매칭.
#   미지원(스킵, content 없음 처리): A4 가로(~841x595pt, 2단인데 폭이 넓어 A방식 크롭 무의미) ·
#     고해상도 스캔 이미지(~2352x3969pt, 텍스트 레이어 없음 — OCR 필요, 범위 밖).
_TEMPLATE_A_MAX_WIDTH = 400
_TEMPLATE_B_WIDTH_RANGE = (560, 620)

_DISASTER_ATTACH_SPLIT_X = 133   # Template A 좌(발생원인)/우(예방대책) 컬럼 경계(254pt 폭 페이지 기준)
# Template A 하단 고정 문구("※본 OPS는 동종재해예방을 목적으로...")가 페이지 폭 전체에 걸쳐 있어
# 좌/우 크롭 경계에서 문장이 반으로 잘린다 — 어느 쪽 크롭에 걸리든 식별 가능한 부분 문자열로 그 줄부터 제거.
# ⚠️ "동종재해예방대책"(Template B 의 실제 예방대책 헤더 표기 중 하나)과 겹치지 않도록 뒤에 "을"을 붙여 한정.
_DISASTER_ATTACH_FOOTER_MARKERS = ("동종재해예방을", "재해발생상황과다를수도있음")

_CAUSE_HEADER_RE = re.compile(r"발생원인")
_MEASURE_HEADER_RE = re.compile(r"예방대책")


def _despace(s):
    return re.sub(r"\s+", "", s or "")


def _find_header_line(lines, pattern, max_header_len=20):
    """짧은 줄(헤더로 추정, 공백 제거 후 20자 이내) 중 패턴이 부분일치하는 첫 줄의 인덱스.
    본문 서술 문장 속에 우연히 같은 문구가 섞여 있어도(문장은 대개 훨씬 길다) 오매칭 방지."""
    for i, ln in enumerate(lines):
        compact = _despace(ln)
        if compact and len(compact) <= max_header_len and pattern.search(compact):
            return i
    return None


def _strip_footer(text):
    lines = text.split("\n")
    for i, ln in enumerate(lines):
        if any(m in ln for m in _DISASTER_ATTACH_FOOTER_MARKERS):
            return "\n".join(lines[:i]).strip()
    return text


def _crop_text(page, x0, x1):
    import warnings
    warnings.filterwarnings("ignore")
    return (page.crop((x0, 0, x1, page.height)).extract_text() or "").strip()


def _after_header(text, header):
    """헤더 단어(발생원인/예방대책) 다음 줄부터만 취한다(그 앞은 재해개요 등 다른 섹션)."""
    lines = text.split("\n")
    idx = next((i for i, ln in enumerate(lines) if header in ln), None)
    if idx is None:
        return ""   # 헤더를 못 찾으면(레이아웃 이례) 빈 문자열 → 이 레코드는 content 없음 처리
    return "\n".join(lines[idx + 1:]).strip()


def _read_disaster_attach_template_a(page):
    """소형 OPS 카드(2026년 최근 케이스): x좌표 크롭으로 좌(발생원인)/우(예방대책) 분리."""
    w = page.width
    left = _crop_text(page, 0, _DISASTER_ATTACH_SPLIT_X)
    right = _crop_text(page, _DISASTER_ATTACH_SPLIT_X, w)
    cause = _strip_footer(_after_header(left, "발생원인")).strip()
    measure = _strip_footer(_after_header(right, "예방대책")).strip()
    return cause, measure


def _read_disaster_attach_template_b(pdf):
    """구형 A4 세로 단일컬럼 서술형 보고서: 전 페이지(최대 3p) 텍스트를 이어붙여 헤더 줄로
    구간 분리. 가로형(2단)은 여기로 오면 좌우 컬럼이 섞여 깨지므로 호출측에서 세로형만 넘긴다."""
    import warnings
    warnings.filterwarnings("ignore")
    all_lines = []
    for page in pdf.pages[:3]:
        all_lines.extend((page.extract_text() or "").split("\n"))
    cause_idx = _find_header_line(all_lines, _CAUSE_HEADER_RE)
    measure_idx = _find_header_line(all_lines, _MEASURE_HEADER_RE)
    if cause_idx is None or measure_idx is None or measure_idx <= cause_idx:
        return "", ""
    cause = "\n".join(all_lines[cause_idx + 1: measure_idx]).strip()
    measure = "\n".join(all_lines[measure_idx + 1:]).strip()
    return cause, measure


def read_disaster_attach_pdf(src):
    """15121008 전용 리더: 15121001 boardno 연계로 수집된 KOSHA 첨부 PDF(재해개요/발생원인/
    예방대책 구성, 템플릿 여러 종 — 위 dispatch 주석 참조)를 파싱한다.
    수집 매니페스트는 20_collect_disaster_attach.py 가 미리 만들어 둔다(PDF 실물도 함께 받아둠).
    레코드 1건 = 첨부 PDF 1건(대다수 boardno당 1개, 일부 다건)."""
    import warnings
    warnings.filterwarnings("ignore")
    import pdfplumber

    sid = str(src["source_id"]).strip()
    manifest_path = os.path.join(RAW_DIR, f"{sid}_kosha_disaster_attach.jsonl")
    if not os.path.exists(manifest_path):
        raise FileNotFoundError(f"수집 매니페스트 없음: {manifest_path} "
                                 f"(20_collect_disaster_attach.py 먼저 실행)")
    manifest_items = []
    with open(manifest_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                manifest_items.append(json.loads(line))

    out = []
    skipped_template = 0
    for it in manifest_items:
        local_path = it.get("local_path")
        if not local_path or not os.path.exists(local_path):
            continue
        try:
            with pdfplumber.open(local_path) as pdf:
                page = pdf.pages[0]
                w, h = page.width, page.height
                if w < _TEMPLATE_A_MAX_WIDTH:
                    cause_text, measure_text = _read_disaster_attach_template_a(page)
                elif _TEMPLATE_B_WIDTH_RANGE[0] <= w <= _TEMPLATE_B_WIDTH_RANGE[1] and h > w:
                    cause_text, measure_text = _read_disaster_attach_template_b(pdf)
                else:
                    cause_text, measure_text = "", ""    # 미지원 템플릿(가로형/스캔 등) — 스킵
                    skipped_template += 1
        except Exception as e:                      # noqa: BLE001 (건 단위 실패는 스킵, 전체는 계속)
            print(f"  [경고] PDF 파싱 실패 {local_path}: {e}")
            continue

        out.append({
            "boardno": it.get("boardno"),
            "filenm": it.get("filenm"),
            "business": it.get("business"),
            "cause_text": cause_text,
            "measure_text": measure_text,
        })
    if skipped_template:
        print(f"  [안내] 미지원 템플릿(가로형/스캔 이미지 등) {skipped_template}건 — content 없음 처리(제외)")
    return out


# 15144147(KOSHA GUIDE) 섹션 헤더 패턴 — "4.1 흙막이 지보공..." / "1. 목적" 형태.
#   TOC 페이지도 같은 번호로 시작하지만 점선 리더(····)+쪽번호가 붙어 있어 그걸로 구분한다.
_KG_SUB_HEADER_RE = re.compile(r"^(\d{1,2}\.\d{1,2}(?:\.\d{1,2})?)\.?\s*[가-힣A-Za-z]")
_KG_TOP_HEADER_RE = re.compile(r"^(\d{1,2})\.\s*[가-힣A-Za-z]")
_KG_TOC_LEADER_RE = re.compile(r"[·.]{3,}")   # TOC 점선 리더 — 있으면 목차 줄로 간주해 헤더 후보에서 제외
_KG_MIN_BODY_LEN = 30                          # 이보다 짧은 섹션은(상위헤더 바로 하위헤더로 이어지는 등) 제외

# ★2026-08-27 사용자 확정: AI EC2 TEI 가 CPU 전용이라 1,039건(23,369청크) 전체 임베딩에 약 19시간
#   소요 — 오늘은 범위를 좁혀 건설(C)·화학공정+건설기술지원규정(D, D-C 하위계열 포함)·기계(M) 만
#   적재한다(측정분석 A계열 등 172~600여 건은 후속 세션으로 미룸). 전량 필요해지면 이 set 을
#   비우거나 확장하면 된다(코드 나머지는 무변경 — 리더 자체는 여전히 전체를 지원).
_KG_ALLOWED_PREFIXES = {"C", "D", "D-C", "M"}


def _kg_running_header_re(tech_gdln_no):
    """페이지마다 반복되는 러닝헤더("KOSHA GUIDE" / "D – C – 1 - 2025" / 쪽번호만 있는 줄)를
    본문에서 제거하기 위한 패턴. 표기 편차(대시 종류·공백)를 흡수하려고 문서번호는 despace 후
    글자 사이 선택적 공백/구두점으로 재구성한다."""
    core = re.sub(r"[\s\-–—]+", "", tech_gdln_no or "")
    if not core:
        return None
    pattern = r"[\s\-–—]*".join(re.escape(ch) for ch in core)
    return re.compile(pattern)


def read_koshaguide_pdf(src):
    """15144147 전용 리더: KOSHA GUIDE PDF(목차 기반 번호 섹션 구조, 1./2./4.1/5.2 등)를
    섹션 단위로 청킹한다. 문서가 12~200p로 길어 통짜 1청크가 아니라 섹션별로 쪼갠다.
    수집 매니페스트는 21_collect_koshaguide_pdf.py 가 미리 만들어 둔다(PDF 실물 포함)."""
    import warnings
    warnings.filterwarnings("ignore")
    import fitz

    sid = str(src["source_id"]).strip()
    manifest_path = os.path.join(RAW_DIR, f"{sid}_koshaguide_pdfs.jsonl")
    if not os.path.exists(manifest_path):
        raise FileNotFoundError(f"수집 매니페스트 없음: {manifest_path} "
                                 f"(21_collect_koshaguide_pdf.py 먼저 실행)")
    manifest_items = []
    with open(manifest_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                manifest_items.append(json.loads(line))

    out = []
    no_sections = 0
    skipped_scope = 0
    for it in manifest_items:
        local_path = it.get("local_path")
        if not local_path or not os.path.exists(local_path):
            continue
        no = it.get("techGdlnNo", "")
        prefix = no.rsplit("-", 2)[0] if "-" in no else no
        if _KG_ALLOWED_PREFIXES and prefix not in _KG_ALLOWED_PREFIXES:
            skipped_scope += 1
            continue
        running_re = _kg_running_header_re(no)
        try:
            doc = fitz.open(local_path)
            lines = []
            for page in doc:
                for ln in page.get_text().split("\n"):
                    s = ln.strip()
                    if not s:
                        continue
                    if running_re and running_re.search(s) and len(s) < 40:
                        continue                        # 러닝헤더/문서코드 줄 제거
                    if s.isdigit() and len(s) <= 3:
                        continue                        # 단독 쪽번호 줄 제거
                    lines.append(s)
            doc.close()
        except Exception as e:                          # noqa: BLE001
            print(f"  [경고] PDF 파싱 실패 {local_path}: {e}")
            continue

        headers = []                                    # (제목, 줄인덱스)
        for i, ln in enumerate(lines):
            if _KG_TOC_LEADER_RE.search(ln):
                continue                                 # 목차 줄(점선 리더) 제외
            m = _KG_SUB_HEADER_RE.match(ln) or _KG_TOP_HEADER_RE.match(ln)
            if m:
                headers.append((ln, i))

        if not headers:
            no_sections += 1
            continue

        for idx, (title, line_idx) in enumerate(headers):
            end_idx = headers[idx + 1][1] if idx + 1 < len(headers) else len(lines)
            body = "\n".join(lines[line_idx + 1: end_idx]).strip()
            if len(body) < _KG_MIN_BODY_LEN:
                continue                                 # 하위헤더로 바로 이어지는 상위헤더 등은 스킵
            out.append({
                "techGdlnNo": no,
                "techGdlnNo_prefix": no.rsplit("-", 2)[0] if "-" in no else no,  # "D-C-1-2025"→"D-C" 등 계열
                "techGdlnNm": it.get("techGdlnNm"),
                "techGdlnOfancYmd": it.get("techGdlnOfancYmd"),
                "section_title": title,
                "section_no": title.split()[0].rstrip("."),
                "section_text": body,
            })
    if no_sections:
        print(f"  [안내] 섹션 헤더 미검출(레이아웃 이례) {no_sections}건 — content 없음 처리(제외)")
    if skipped_scope:
        print(f"  [안내] 이번 라운드 범위 제외({sorted(_KG_ALLOWED_PREFIXES)} 계열만) {skipped_scope}건")
    return out


def read_records(src):
    sid = str(src.get("source_id", "")).strip()
    # 15121008: registry상 feed_type=API 이지만 실제로는 15121001 boardno 연계 + PDF 첨부라
    #   범용 API 리더(read_api, collect_api.py 스냅샷 전제)와 다른 전용 리더를 먼저 태운다.
    if sid == "15121008":
        return read_disaster_attach_pdf(src)
    # 15144147: registry상 feed_type=API 이지만 목록 API는 메타데이터만 주고 실제 본문은 PDF다.
    if sid == "15144147":
        return read_koshaguide_pdf(src)
    feed = str(src.get("feed_type", "")).strip().upper()
    if feed == "API":
        return read_api(src)
    ext = os.path.splitext(os.path.basename(str(src.get("file", "")).strip()))[1].lower()
    if feed == "FILE" and ext == ".csv":
        return read_csv_file(src)
    if feed == "FILE" and ext == ".xlsx" and sid == "15140383":
        return read_sif_xlsx(src)
    if feed == "FILE" and ext == ".pdf" and sid == "15140227":
        return read_pdf(src)
    raise NotImplementedError(f"{sid}: feed_type={feed}, ext={ext} 리더 미구현")


# ── 변환: 레코드 → 청크직전 dict ──
def transform(src, rec):
    sid = str(src["source_id"]).strip()
    content_fields = src.get("content_fields", [])
    mask_fields = src.get("mask_fields", [])
    meta_fields = src.get("meta_fields", [])
    excl = src.get("exclude_from_llm", [])
    locator_fields = src.get("locator_fields", [])
    hazard_field = str(src.get("hazard_field", "")).strip()
    measure_field = str(src.get("measure_field", "")).strip()
    domain_field = str(src.get("domain_field", "")).strip()
    cause_field = str(src.get("cause_agent_field", "")).strip()

    # CONTENT 조립
    parts = [normalize_text(rec.get(f)) for f in content_fields]
    content = "\n\n".join(p for p in parts if p)
    if not content:
        return None  # content 없음 → 행 제외(§4-5)

    # 마스킹(결정③): 주소는 시군구 절단, 그 외 식별값은 치환, 업체/성명 종합 마스킹
    addr_vals = [rec.get(f) for f in mask_fields if is_addr_field(f)]
    other_vals = [rec.get(f) for f in mask_fields if not is_addr_field(f)]
    content = _mask.mask_body(content, addr_vals, other_vals)

    # HAZARD/MEASURE
    hazard_text = _mask.mask_body(normalize_text(rec.get(hazard_field)), addr_vals, other_vals) if hazard_field else None
    hazard_text = hazard_text or None
    measure_text = measure_or_null(rec.get(measure_field)) if measure_field else None
    if measure_text:
        measure_text = _mask.mask_body(measure_text, addr_vals, other_vals)

    # META_JSON = meta_fields + exclude_from_llm(감사용). 주소는 시군구 절단, 그 외 mask_field는 익명치환.
    meta = {}
    for f in list(meta_fields) + list(excl):
        meta[f] = rec.get(f)
    for f in mask_fields:
        v = rec.get(f)
        if is_addr_field(f):
            meta[f] = _mask.truncate_address(v)
        else:
            # ★2026-08-26: mask_company/mask_person 패턴매칭에 맡기지 않는다 — 15069200 실측에서
            # 건설업 접미사가 아닌 화학/제조업 회사명(케미칼/켐텍/LED/브랜드명 등)이 패턴에 안 걸려
            # 실명이 그대로 남는 사고 발견. mask_fields 에 오른 시점에 이미 "식별값"임이 확정이므로
            # 패턴으로 재확인(추측)하지 않고 무조건 치환한다(하드가드#2: 애매하면 제거).
            meta[f] = "○○" if v else v

    locator = "|".join(str(rec.get(f, "")).strip() for f in locator_fields)
    return {
        "chunk_id": chunk_id(sid, locator, content),
        "source_id": sid,
        "content": content,
        "hazard_text": hazard_text,
        "measure_text": measure_text,
        "domain_tag": (normalize_text(rec.get(domain_field)) or None) if domain_field else None,
        "cause_agent": (normalize_text(rec.get(cause_field)) or None) if cause_field else None,
        "meta_json": meta,
        "source_locator": locator or None,
        # 소스 레벨 메타(하드가드#4 verbatim 게이팅·신뢰등급 표시가 청크까지 따라가게)
        "track": str(src.get("track", "")).strip() or None,
        "data_reliability": str(src.get("data_reliability", "")).strip() or None,
        "source_name": str(src.get("source_name", "")).strip() or None,
        "use_yn": "Y",
    }


def process_source(src):
    sid = str(src["source_id"]).strip()
    records = read_records(src)
    out, excluded, seen = [], 0, set()
    for rec in records:
        row = transform(src, rec)
        if row is None:
            excluded += 1
            continue
        if row["chunk_id"] in seen:      # 완전 동일 CONTENT → 멱등 dedup
            continue
        seen.add(row["chunk_id"])
        out.append(row)

    os.makedirs(PROCESSED_DIR, exist_ok=True)
    out_path = os.path.join(PROCESSED_DIR, f"{sid}_processed.json")
    json.dump(out, open(out_path, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    n_meas = sum(1 for r in out if r["measure_text"])
    n_haz = sum(1 for r in out if r["hazard_text"])
    print(f"[{sid}] 입력 {len(records)} · 제외 {excluded} · 청크 {len(out)} "
          f"(measure {n_meas}, hazard {n_haz}) → {os.path.basename(out_path)}")
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", help="특정 source_id만")
    args = ap.parse_args()

    reg = _ca.load_registry(REGISTRY_PATH)
    targets = [s for s in reg if _ca.is_checked(s.get("license_checked"))
               and (not args.source or str(s.get("source_id")) == str(args.source))]
    if not targets:
        sys.exit("[안내] 처리대상(license_checked=TRUE) 없음.")

    print(f"== 전처리 시작 · 대상 {len(targets)}건 ==")
    for src in targets:
        try:
            process_source(src)
        except NotImplementedError as e:
            print(f"[{src['source_id']}] 건너뜀 — {e}")
        except Exception as e:      # noqa: BLE001
            print(f"[{src['source_id']}] 오류 — {e}")


if __name__ == "__main__":
    main()
